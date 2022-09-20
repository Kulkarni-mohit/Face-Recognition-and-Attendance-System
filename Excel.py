
from openpyxl import load_workbook
import time


inclass = {}
def excel(name):
    
    
    workbook = load_workbook(filename="Attendance file.xlsx")
    sheet = workbook.active
    
    t = time.localtime()

    list_stu = [['Mohit_Kulkarni',203,7499986979],['Kapil_Kulkarni',205,9876543221],['Harshal_Korade',202,879654123]]

    cur = sheet.max_row
    
    if (name in inclass) and (inclass[name][1]- time.time()>5):
        sheet['F'+str(inclass[name])]=time.strftime("%H:%M:%S", t)
        del inclass[name]

    elif name not in inclass:
        for i in list_stu:
            
            if name == str(i[0]):
                sheet["B"+str(cur+1)] = i[0]
                sheet["C"+str(cur+1)] = i[1]
                sheet["D"+str(cur+1)] = i[2]
                sheet["E"+str(cur+1)] = time.strftime("%H:%M:%S", t)
                inclass[name]=[cur+1,time.time()]
    workbook.save(filename="Attendance file.xlsx")
    
    print(inclass)
